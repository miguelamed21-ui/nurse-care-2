"""
Utilidades para exportación de datos en diferentes formatos (PDF, Excel)
para la plataforma AMED-IA
"""

from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import RadarChart, Reference
from datetime import datetime
from io import BytesIO
import matplotlib
matplotlib.use('Agg')  # Non-interactive backend
import matplotlib.pyplot as plt
import numpy as np

# Paleta de colores AMED-IA
COLORS = {
    'primary': '#005A9C',
    'secondary': '#10B981',
    'accent': '#3B82F6',
    'text_dark': '#334155',
    'text_light': '#64748B',
    'background': '#F8FAFC',
    'border': '#E2E8F0'
}

# Nombres de competencias
COMPETENCY_NAMES = {
    'comunicacion': 'Comunicación',
    'valoracion_clinica': 'Valoración Clínica',
    'razonamiento_critico': 'Razonamiento Crítico',
    'competencia_tecnica': 'Competencia Técnica',
    'empatia': 'Empatía y Relación Terapéutica'
}

def generate_evaluation_pdf(evaluation_data, case_data, user_data):
    """
    Genera PDF de evaluación de caso individual
    
    Args:
        evaluation_data: dict con eval_id, scores, feedback, strengths, improvements, evaluated_at
        case_data: dict con title, specialty, difficulty
        user_data: dict con name, email
    
    Returns:
        BytesIO object con el PDF generado
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=0.75*inch, leftMargin=0.75*inch,
                            topMargin=1*inch, bottomMargin=0.75*inch)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Estilos personalizados
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor(COLORS['primary']),
        spaceAfter=12,
        alignment=TA_CENTER
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor(COLORS['primary']),
        spaceBefore=12,
        spaceAfter=6
    )
    
    body_style = ParagraphStyle(
        'CustomBody',
        parent=styles['BodyText'],
        fontSize=10,
        textColor=colors.HexColor(COLORS['text_dark'])
    )
    
    # Header con logo (texto por ahora)
    header = Paragraph("AMED-IA", title_style)
    elements.append(header)
    
    subtitle = Paragraph("Evaluación de Competencias en Enfermería", body_style)
    elements.append(subtitle)
    elements.append(Spacer(1, 0.3*inch))
    
    # Línea decorativa
    line_data = [[''] * 1]
    line_table = Table(line_data, colWidths=[6.5*inch])
    line_table.setStyle(TableStyle([
        ('LINEABOVE', (0, 0), (-1, 0), 2, colors.HexColor(COLORS['primary'])),
    ]))
    elements.append(line_table)
    elements.append(Spacer(1, 0.2*inch))
    
    # Información del documento
    elements.append(Paragraph("Evaluación de Caso Individual", heading_style))
    
    # Tabla de información general
    eval_date = evaluation_data.get('evaluated_at', datetime.now())
    if isinstance(eval_date, str):
        eval_date = datetime.fromisoformat(eval_date.replace('Z', '+00:00'))
    
    info_data = [
        ['Estudiante:', user_data.get('name', 'N/A')],
        ['Caso Clínico:', case_data.get('title', 'N/A')],
        ['Especialidad:', case_data.get('specialty', 'N/A')],
        ['Dificultad:', case_data.get('difficulty', 'N/A')],
        ['Fecha de Evaluación:', eval_date.strftime('%d/%m/%Y %H:%M')],
        ['ID de Evaluación:', evaluation_data.get('eval_id', 'N/A')[:12]]
    ]
    
    info_table = Table(info_data, colWidths=[2*inch, 4.5*inch])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor(COLORS['background'])),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor(COLORS['text_dark'])),
        ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor(COLORS['border'])),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Scores por competencia
    elements.append(Paragraph("Resultados por Competencia", heading_style))
    
    scores = evaluation_data.get('scores', {})
    score_data = [['Competencia', 'Puntaje', 'Nivel de Logro']]
    
    for key, value in scores.items():
        comp_name = COMPETENCY_NAMES.get(key, key.replace('_', ' ').title())
        nivel = ''
        if value <= 25:
            nivel = 'Inicial / Deficiente (0-25%)'
        elif value <= 50:
            nivel = 'En Desarrollo (26-50%)'
        elif value <= 75:
            nivel = 'Competente (51-75%)'
        else:
            nivel = 'Excelente (76-100%)'
        
        score_data.append([comp_name, f"{value}%", nivel])
    
    # Calcular promedio
    avg_score = sum(scores.values()) / len(scores) if scores else 0
    score_data.append(['PROMEDIO GENERAL', f"{avg_score:.1f}%", ''])
    
    score_table = Table(score_data, colWidths=[2.5*inch, 1.5*inch, 2.5*inch])
    score_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(COLORS['primary'])),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -2), 10),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor(COLORS['accent'])),
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.whitesmoke),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor(COLORS['border'])),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(score_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Feedback
    elements.append(Paragraph("Retroalimentación General", heading_style))
    feedback_text = evaluation_data.get('feedback', 'No se proporcionó retroalimentación.')
    elements.append(Paragraph(feedback_text, body_style))
    elements.append(Spacer(1, 0.2*inch))
    
    # Fortalezas
    strengths = evaluation_data.get('strengths', [])
    if strengths and len(strengths) > 0:
        elements.append(Paragraph("Fortalezas Observadas", heading_style))
        for strength in strengths:
            if strength.strip():
                elements.append(Paragraph(f"• {strength}", body_style))
        elements.append(Spacer(1, 0.2*inch))
    
    # Áreas de mejora
    improvements = evaluation_data.get('improvements', [])
    if improvements and len(improvements) > 0:
        elements.append(Paragraph("Áreas de Mejora", heading_style))
        for improvement in improvements:
            if improvement.strip():
                elements.append(Paragraph(f"• {improvement}", body_style))
        elements.append(Spacer(1, 0.2*inch))
    
    # Footer
    elements.append(Spacer(1, 0.5*inch))
    footer_text = f"Generado por AMED-IA el {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.HexColor(COLORS['text_light']),
        alignment=TA_CENTER
    )
    elements.append(Paragraph(footer_text, footer_style))
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    return buffer


def generate_evaluation_excel(evaluation_data, case_data, user_data):
    """
    Genera Excel de evaluación de caso individual
    
    Args:
        evaluation_data: dict con eval_id, scores, feedback, strengths, improvements, evaluated_at
        case_data: dict con title, specialty, difficulty
        user_data: dict con name, email
    
    Returns:
        BytesIO object con el Excel generado
    """
    buffer = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Evaluación"
    
    # Estilos
    header_fill = PatternFill(start_color=COLORS['primary'].replace('#', ''), end_color=COLORS['primary'].replace('#', ''), fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)
    title_font = Font(bold=True, size=16, color=COLORS['primary'].replace('#', ''))
    normal_font = Font(size=11)
    bold_font = Font(bold=True, size=11)
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título
    ws['A1'] = 'AMED-IA - Evaluación de Competencias en Enfermería'
    ws['A1'].font = title_font
    ws.merge_cells('A1:D1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Información general
    ws['A3'] = 'Evaluación de Caso Individual'
    ws['A3'].font = Font(bold=True, size=14)
    
    eval_date = evaluation_data.get('evaluated_at', datetime.now())
    if isinstance(eval_date, str):
        eval_date = datetime.fromisoformat(eval_date.replace('Z', '+00:00'))
    
    info_rows = [
        ('Estudiante:', user_data.get('name', 'N/A')),
        ('Caso Clínico:', case_data.get('title', 'N/A')),
        ('Especialidad:', case_data.get('specialty', 'N/A')),
        ('Dificultad:', case_data.get('difficulty', 'N/A')),
        ('Fecha de Evaluación:', eval_date.strftime('%d/%m/%Y %H:%M')),
        ('ID de Evaluación:', evaluation_data.get('eval_id', 'N/A')[:12])
    ]
    
    row = 5
    for label, value in info_rows:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = bold_font
        ws[f'B{row}'] = value
        ws[f'B{row}'].font = normal_font
        row += 1
    
    # Scores
    row += 2
    ws[f'A{row}'] = 'Resultados por Competencia'
    ws[f'A{row}'].font = Font(bold=True, size=12)
    
    row += 1
    headers = ['Competencia', 'Puntaje', 'Nivel de Logro']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    scores = evaluation_data.get('scores', {})
    row += 1
    for key, value in scores.items():
        comp_name = COMPETENCY_NAMES.get(key, key.replace('_', ' ').title())
        nivel = ''
        if value <= 25:
            nivel = 'Inicial / Deficiente (0-25%)'
        elif value <= 50:
            nivel = 'En Desarrollo (26-50%)'
        elif value <= 75:
            nivel = 'Competente (51-75%)'
        else:
            nivel = 'Excelente (76-100%)'
        
        ws[f'A{row}'] = comp_name
        ws[f'B{row}'] = f"{value}%"
        ws[f'C{row}'] = nivel
        
        for col in range(1, 4):
            ws.cell(row=row, column=col).border = border
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='left', vertical='center')
        
        row += 1
    
    # Promedio
    avg_score = sum(scores.values()) / len(scores) if scores else 0
    ws[f'A{row}'] = 'PROMEDIO GENERAL'
    ws[f'A{row}'].font = bold_font
    ws[f'B{row}'] = f"{avg_score:.1f}%"
    ws[f'B{row}'].font = bold_font
    for col in range(1, 4):
        ws.cell(row=row, column=col).border = border
        ws.cell(row=row, column=col).fill = PatternFill(start_color=COLORS['accent'].replace('#', ''), end_color=COLORS['accent'].replace('#', ''), fill_type='solid')
        ws.cell(row=row, column=col).font = Font(bold=True, color='FFFFFF')
    
    # Feedback
    row += 3
    ws[f'A{row}'] = 'Retroalimentación General'
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1
    ws[f'A{row}'] = evaluation_data.get('feedback', 'No se proporcionó retroalimentación.')
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
    
    # Fortalezas
    strengths = evaluation_data.get('strengths', [])
    if strengths and len(strengths) > 0:
        row += 3
        ws[f'A{row}'] = 'Fortalezas Observadas'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        for strength in strengths:
            if strength.strip():
                row += 1
                ws[f'A{row}'] = f"• {strength}"
    
    # Áreas de mejora
    improvements = evaluation_data.get('improvements', [])
    if improvements and len(improvements) > 0:
        row += 3
        ws[f'A{row}'] = 'Áreas de Mejora'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        for improvement in improvements:
            if improvement.strip():
                row += 1
                ws[f'A{row}'] = f"• {improvement}"
    
    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 20
    
    # Guardar
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_global_evaluation_pdf(evaluation_data, user_data, simulations_data):
    """
    Genera PDF de evaluación global/transversal
    
    Args:
        evaluation_data: dict con eval_id, scores, feedback, period_info, evaluated_at
        user_data: dict con name, email
        simulations_data: list de simulaciones incluidas
    
    Returns:
        BytesIO object con el PDF generado
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=0.75*inch, leftMargin=0.75*inch,
                            topMargin=1*inch, bottomMargin=0.75*inch)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Estilos personalizados
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor(COLORS['primary']),
        spaceAfter=12,
        alignment=TA_CENTER
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor(COLORS['primary']),
        spaceBefore=12,
        spaceAfter=6
    )
    
    body_style = ParagraphStyle(
        'CustomBody',
        parent=styles['BodyText'],
        fontSize=10,
        textColor=colors.HexColor(COLORS['text_dark'])
    )
    
    # Header
    header = Paragraph("AMED-IA", title_style)
    elements.append(header)
    
    subtitle = Paragraph("Evaluación Global de Competencias", body_style)
    elements.append(subtitle)
    elements.append(Spacer(1, 0.3*inch))
    
    # Línea decorativa
    line_data = [[''] * 1]
    line_table = Table(line_data, colWidths=[6.5*inch])
    line_table.setStyle(TableStyle([
        ('LINEABOVE', (0, 0), (-1, 0), 2, colors.HexColor(COLORS['primary'])),
    ]))
    elements.append(line_table)
    elements.append(Spacer(1, 0.2*inch))
    
    # Información del documento
    elements.append(Paragraph("Evaluación Global / Transversal", heading_style))
    
    # Tabla de información general
    eval_date = evaluation_data.get('evaluated_at', datetime.now())
    if isinstance(eval_date, str):
        eval_date = datetime.fromisoformat(eval_date.replace('Z', '+00:00'))
    
    period_info = evaluation_data.get('period_info', {})
    start_date = period_info.get('start_date', 'N/A')
    end_date = period_info.get('end_date', 'N/A')
    sims_count = period_info.get('simulations_count', len(simulations_data))
    
    info_data = [
        ['Estudiante:', user_data.get('name', 'N/A')],
        ['Periodo:', f"{start_date} a {end_date}" if start_date != 'N/A' else 'Todos los registros'],
        ['Simulaciones Evaluadas:', str(sims_count)],
        ['Fecha de Generación:', eval_date.strftime('%d/%m/%Y %H:%M')],
        ['ID de Evaluación:', evaluation_data.get('eval_id', 'N/A')[:12]]
    ]
    
    info_table = Table(info_data, colWidths=[2.5*inch, 4*inch])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor(COLORS['background'])),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor(COLORS['text_dark'])),
        ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor(COLORS['border'])),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Promedio Global
    elements.append(Paragraph("Promedio Global de Competencias", heading_style))
    
    scores = evaluation_data.get('scores', {})
    score_data = [['Competencia', 'Promedio', 'Nivel de Logro']]
    
    total_score = 0
    for key, value in scores.items():
        comp_name = COMPETENCY_NAMES.get(key, key.replace('_', ' ').title())
        nivel = ''
        if value <= 25:
            nivel = 'Inicial / Deficiente (0-25%)'
        elif value <= 50:
            nivel = 'En Desarrollo (26-50%)'
        elif value <= 75:
            nivel = 'Competente (51-75%)'
        else:
            nivel = 'Excelente (76-100%)'
        
        score_data.append([comp_name, f"{value:.1f}%", nivel])
        total_score += value
    
    # Promedio general
    avg_score = total_score / len(scores) if scores else 0
    score_data.append(['PROMEDIO GENERAL', f"{avg_score:.1f}%", ''])
    
    score_table = Table(score_data, colWidths=[2.5*inch, 1.5*inch, 2.5*inch])
    score_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(COLORS['secondary'])),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -2), 10),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor(COLORS['accent'])),
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.whitesmoke),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor(COLORS['border'])),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(score_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Lista de casos evaluados
    if simulations_data and len(simulations_data) > 0:
        elements.append(Paragraph("Casos Evaluados en este Periodo", heading_style))
        
        cases_data = [['#', 'Caso', 'Fecha', 'Estado']]
        for idx, sim in enumerate(simulations_data[:10], 1):  # Max 10 casos en PDF
            case_title = sim.get('case_title', 'Caso no disponible')[:40]
            started_at = sim.get('started_at', '')
            if started_at:
                started_at = datetime.fromisoformat(started_at.replace('Z', '+00:00')).strftime('%d/%m/%Y')
            status = 'Completado' if sim.get('status') == 'completed' else sim.get('status', 'N/A')
            cases_data.append([str(idx), case_title, started_at, status])
        
        if len(simulations_data) > 10:
            cases_data.append(['...', f'Y {len(simulations_data) - 10} más', '', ''])
        
        cases_table = Table(cases_data, colWidths=[0.5*inch, 3.5*inch, 1.5*inch, 1*inch])
        cases_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(COLORS['background'])),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor(COLORS['text_dark'])),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor(COLORS['border'])),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(cases_table)
        elements.append(Spacer(1, 0.2*inch))
    
    # Feedback
    feedback = evaluation_data.get('feedback', '')
    if feedback:
        elements.append(Paragraph("Análisis del Periodo", heading_style))
        elements.append(Paragraph(feedback, body_style))
    
    # Footer
    elements.append(Spacer(1, 0.5*inch))
    footer_text = f"Generado por AMED-IA el {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.HexColor(COLORS['text_light']),
        alignment=TA_CENTER
    )
    elements.append(Paragraph(footer_text, footer_style))
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    return buffer


def generate_global_evaluation_excel(evaluation_data, user_data, simulations_data):
    """
    Genera Excel de evaluación global/transversal
    
    Args:
        evaluation_data: dict con eval_id, scores, feedback, period_info, evaluated_at
        user_data: dict con name, email
        simulations_data: list de simulaciones incluidas
    
    Returns:
        BytesIO object con el Excel generado
    """
    buffer = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Evaluación Global"
    
    # Estilos
    header_fill = PatternFill(start_color=COLORS['secondary'].replace('#', ''), end_color=COLORS['secondary'].replace('#', ''), fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)
    title_font = Font(bold=True, size=16, color=COLORS['primary'].replace('#', ''))
    bold_font = Font(bold=True, size=11)
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título
    ws['A1'] = 'AMED-IA - Evaluación Global de Competencias'
    ws['A1'].font = title_font
    ws.merge_cells('A1:D1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Información general
    ws['A3'] = 'Evaluación Global / Transversal'
    ws['A3'].font = Font(bold=True, size=14)
    
    eval_date = evaluation_data.get('evaluated_at', datetime.now())
    if isinstance(eval_date, str):
        eval_date = datetime.fromisoformat(eval_date.replace('Z', '+00:00'))
    
    period_info = evaluation_data.get('period_info', {})
    start_date = period_info.get('start_date', 'N/A')
    end_date = period_info.get('end_date', 'N/A')
    sims_count = period_info.get('simulations_count', len(simulations_data))
    
    info_rows = [
        ('Estudiante:', user_data.get('name', 'N/A')),
        ('Periodo:', f"{start_date} a {end_date}" if start_date != 'N/A' else 'Todos los registros'),
        ('Simulaciones Evaluadas:', str(sims_count)),
        ('Fecha de Generación:', eval_date.strftime('%d/%m/%Y %H:%M')),
        ('ID de Evaluación:', evaluation_data.get('eval_id', 'N/A')[:12])
    ]
    
    row = 5
    for label, value in info_rows:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = bold_font
        ws[f'B{row}'] = value
        row += 1
    
    # Scores
    row += 2
    ws[f'A{row}'] = 'Promedio Global de Competencias'
    ws[f'A{row}'].font = Font(bold=True, size=12)
    
    row += 1
    headers = ['Competencia', 'Promedio', 'Nivel de Logro']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    scores = evaluation_data.get('scores', {})
    total_score = 0
    row += 1
    for key, value in scores.items():
        comp_name = COMPETENCY_NAMES.get(key, key.replace('_', ' ').title())
        nivel = ''
        if value <= 25:
            nivel = 'Inicial / Deficiente (0-25%)'
        elif value <= 50:
            nivel = 'En Desarrollo (26-50%)'
        elif value <= 75:
            nivel = 'Competente (51-75%)'
        else:
            nivel = 'Excelente (76-100%)'
        
        ws[f'A{row}'] = comp_name
        ws[f'B{row}'] = f"{value:.1f}%"
        ws[f'C{row}'] = nivel
        
        for col in range(1, 4):
            ws.cell(row=row, column=col).border = border
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='left', vertical='center')
        
        total_score += value
        row += 1
    
    # Promedio general
    avg_score = total_score / len(scores) if scores else 0
    ws[f'A{row}'] = 'PROMEDIO GENERAL'
    ws[f'A{row}'].font = bold_font
    ws[f'B{row}'] = f"{avg_score:.1f}%"
    ws[f'B{row}'].font = bold_font
    for col in range(1, 4):
        ws.cell(row=row, column=col).border = border
        ws.cell(row=row, column=col).fill = PatternFill(start_color=COLORS['accent'].replace('#', ''), end_color=COLORS['accent'].replace('#', ''), fill_type='solid')
        ws.cell(row=row, column=col).font = Font(bold=True, color='FFFFFF')
    
    # Lista de casos
    if simulations_data:
        row += 3
        ws[f'A{row}'] = 'Casos Evaluados en este Periodo'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        
        row += 1
        case_headers = ['#', 'Caso', 'Fecha', 'Estado']
        for col, header in enumerate(case_headers, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        for idx, sim in enumerate(simulations_data, 1):
            row += 1
            ws[f'A{row}'] = idx
            ws[f'B{row}'] = sim.get('case_title', 'Caso no disponible')[:50]
            started_at = sim.get('started_at', '')
            if started_at:
                started_at = datetime.fromisoformat(started_at.replace('Z', '+00:00')).strftime('%d/%m/%Y')
            ws[f'C{row}'] = started_at
            ws[f'D{row}'] = 'Completado' if sim.get('status') == 'completed' else sim.get('status', 'N/A')
            
            for col in range(1, 5):
                ws.cell(row=row, column=col).border = border
    
    # Ajustar anchos
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def generate_group_analytics_pdf(group_data, students_data, stats):
    """
    Genera PDF con analíticas de grupo
    
    Args:
        group_data: dict con group_name, teacher_name
        students_data: list de estudiantes con sus métricas
        stats: dict con métricas agregadas del grupo
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#005A9C'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    elements.append(Paragraph(f"Analíticas de Grupo: {group_data['group_name']}", title_style))
    elements.append(Spacer(1, 0.3*inch))
    
    # Metadata
    meta_style = ParagraphStyle('Meta', parent=styles['Normal'], fontSize=10, textColor=colors.HexColor('#64748B'))
    elements.append(Paragraph(f"<b>Docente:</b> {group_data.get('teacher_name', 'N/A')}", meta_style))
    elements.append(Paragraph(f"<b>Fecha:</b> {datetime.now().strftime('%d/%m/%Y %H:%M')}", meta_style))
    elements.append(Paragraph(f"<b>Total Estudiantes:</b> {stats['student_count']}", meta_style))
    elements.append(Spacer(1, 0.3*inch))
    
    # Summary Stats
    elements.append(Paragraph("<b>Resumen General</b>", styles['Heading2']))
    summary_data = [
        ['Métrica', 'Valor'],
        ['Simulaciones Totales', str(stats['total_simulations'])],
        ['Simulaciones Completadas', str(stats['completed_simulations'])],
        ['Evaluaciones Totales', str(stats['total_evaluations'])],
        ['IDEC Promedio', f"{stats['avg_idec']:.2f}"]
    ]
    
    summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#005A9C')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Competency Averages
    elements.append(Paragraph("<b>Competencias Promedio</b>", styles['Heading2']))
    comp_data = [['Competencia', 'Puntuación']]
    for comp, score in stats.get('avg_competencies', {}).items():
        comp_data.append([comp, f"{score:.2f}"])
    
    comp_table = Table(comp_data, colWidths=[4*inch, 1.5*inch])
    comp_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10B981')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(comp_table)
    elements.append(PageBreak())
    
    # Students Table
    elements.append(Paragraph("<b>Resumen por Estudiante</b>", styles['Heading2']))
    student_data = [['Estudiante', 'Sims', 'Evals', 'IDEC']]
    for student in students_data[:20]:  # Limit to 20 students per page
        student_data.append([
            student['name'][:30],
            str(student['simulations_count']),
            str(student['evaluations_count']),
            f"{student['idec_score']:.1f}"
        ])
    
    student_table = Table(student_data, colWidths=[3.5*inch, 0.8*inch, 0.8*inch, 0.8*inch])
    student_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#8B5CF6')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(student_table)
    
    doc.build(elements)
    buffer.seek(0)
    return buffer


def generate_group_comparison_excel(group_name, comparisons):
    """
    Genera Excel con comparativa de estudiantes del grupo
    
    Args:
        group_name: nombre del grupo
        comparisons: lista de estudiantes con todas sus métricas
    """
    buffer = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparativa"
    
    # Header styling
    header_fill = PatternFill(start_color="005A9C", end_color="005A9C", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws['A1'] = f"Comparativa de Estudiantes - {group_name}"
    ws['A1'].font = Font(size=16, bold=True, color="005A9C")
    ws.merge_cells('A1:H1')
    
    ws['A2'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A2'].font = Font(size=10, color="64748B")
    ws.merge_cells('A2:H2')
    
    # Headers
    headers = ['Nombre', 'Email', 'Simulaciones', 'Completadas', 'Evaluaciones', 'IDEC', 'Ranking']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Add competency headers
    from constants import COMPETENCY_NAMES
    col_offset = len(headers) + 1
    for idx, comp in enumerate(COMPETENCY_NAMES, start=col_offset):
        cell = ws.cell(row=4, column=idx, value=comp[:20])
        cell.fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True, size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    # Data rows
    for idx, student in enumerate(comparisons, start=5):
        ws.cell(row=idx, column=1, value=student['name'])
        ws.cell(row=idx, column=2, value=student['email'])
        ws.cell(row=idx, column=3, value=student['simulations'])
        ws.cell(row=idx, column=4, value=student['completed_simulations'])
        ws.cell(row=idx, column=5, value=student['evaluations'])
        ws.cell(row=idx, column=6, value=round(student['idec_score'], 2))
        ws.cell(row=idx, column=7, value=idx - 4)  # Ranking
        
        # Competency scores
        for comp_idx, comp_name in enumerate(COMPETENCY_NAMES, start=col_offset):
            score = student.get('competencies', {}).get(comp_name, 0)
            ws.cell(row=idx, column=comp_idx, value=round(score, 2))
        
        # Apply borders
        for col in range(1, col_offset + len(COMPETENCY_NAMES)):
            ws.cell(row=idx, column=col).border = border
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 10
    
    for idx in range(col_offset, col_offset + len(COMPETENCY_NAMES)):
        ws.column_dimensions[chr(64 + idx)].width = 18
    
    wb.save(buffer)
    buffer.seek(0)
    return buffer
